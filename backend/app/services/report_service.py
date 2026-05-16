"""
Generates CSV and Excel reports required by BRD:
  - Achievement Report: Planned Target vs Actual for all employees
  - Completion Dashboard: who has/hasn't completed check-ins
"""
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from ..models.goal import GoalSheet, Goal, QuarterlyActual, Cycle, SheetStatus
from ..models.user import User


def _get_all_sheets(cycle_id: int, db: Session):
    return (
        db.query(GoalSheet)
        .filter(GoalSheet.cycle_id == cycle_id)
        .join(User, User.id == GoalSheet.employee_id)
        .all()
    )


# ── CSV Export ────────────────────────────────────────────────────────────────

def generate_achievement_csv(cycle_id: int, db: Session) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "Employee ID", "Employee Name", "Department", "Manager",
        "Goal Title", "Thrust Area", "UoM Type",
        "Target", "Actual", "Progress Score (%)", "Status",
        "Sheet Status", "Approved At",
    ])

    sheets = _get_all_sheets(cycle_id, db)
    for sheet in sheets:
        employee = sheet.employee
        manager_name = employee.manager.name if employee.manager else "—"

        for goal in sheet.goals:
            actuals = [a for a in goal.quarterly_actuals if a.cycle_id == cycle_id]
            if actuals:
                actual = actuals[0]
                actual_val = actual.actual_numeric or str(actual.actual_date or "")
                score = f"{actual.progress_score:.1f}" if actual.progress_score is not None else "—"
                g_status = actual.status.value
            else:
                actual_val = "—"
                score = "—"
                g_status = "not_started"

            target_val = goal.target_numeric if goal.target_numeric is not None else str(goal.target_date or "")

            writer.writerow([
                employee.id, employee.name, employee.department or "—", manager_name,
                goal.title,
                goal.thrust_area.name if goal.thrust_area else "—",
                goal.uom_type.value,
                target_val, actual_val, score, g_status,
                sheet.status.value,
                sheet.approved_at.strftime("%Y-%m-%d") if sheet.approved_at else "—",
            ])

    return output.getvalue().encode("utf-8")


# ── Excel Export ──────────────────────────────────────────────────────────────

def _header_style(ws, headers: list, row: int = 1):
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = max(len(header) + 4, 14)

    ws.row_dimensions[row].height = 30


def _data_style(ws, row: int, col: int, value, score=False):
    thin = Side(style="thin", color="EEEEEE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = border
    cell.alignment = Alignment(vertical="center", wrap_text=True)

    if score and isinstance(value, (int, float)):
        if value >= 90:
            cell.fill = PatternFill("solid", fgColor="C6EFCE")
            cell.font = Font(color="276221")
        elif value >= 70:
            cell.fill = PatternFill("solid", fgColor="FFEB9C")
            cell.font = Font(color="9C5700")
        else:
            cell.fill = PatternFill("solid", fgColor="FFC7CE")
            cell.font = Font(color="9C0006")
    return cell


def generate_achievement_excel(cycle_id: int, db: Session) -> bytes:
    wb = Workbook()
    ws = wb.active
    cycle = db.query(Cycle).filter(Cycle.id == cycle_id).first()
    ws.title = f"Achievement Report"

    # Title row
    ws.merge_cells("A1:M1")
    title_cell = ws["A1"]
    cycle_label = f"{cycle.year} — {cycle.phase.value.upper()}" if cycle else f"Cycle {cycle_id}"
    title_cell.value = f"AtomQuest — Achievement Report | {cycle_label}"
    title_cell.font = Font(bold=True, size=14, color="1F3864")
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 35

    headers = [
        "Emp ID", "Employee Name", "Department", "Manager",
        "Goal Title", "Thrust Area", "UoM Type",
        "Target", "Actual", "Score (%)", "Goal Status",
        "Sheet Status", "Approved At",
    ]
    _header_style(ws, headers, row=2)

    sheets = _get_all_sheets(cycle_id, db)
    row = 3
    for sheet in sheets:
        employee = sheet.employee
        manager_name = employee.manager.name if employee.manager else "—"

        for goal in sheet.goals:
            actuals = [a for a in goal.quarterly_actuals if a.cycle_id == cycle_id]
            if actuals:
                actual = actuals[0]
                actual_val = actual.actual_numeric if actual.actual_numeric is not None else str(actual.actual_date or "")
                score_val = round(actual.progress_score, 1) if actual.progress_score is not None else None
                g_status = actual.status.value
            else:
                actual_val = "—"
                score_val = None
                g_status = "not_started"

            target_val = goal.target_numeric if goal.target_numeric is not None else str(goal.target_date or "")

            values = [
                employee.id, employee.name, employee.department or "—", manager_name,
                goal.title,
                goal.thrust_area.name if goal.thrust_area else "—",
                goal.uom_type.value,
                target_val, actual_val,
                score_val if score_val is not None else "—",
                g_status, sheet.status.value,
                sheet.approved_at.strftime("%Y-%m-%d") if sheet.approved_at else "—",
            ]

            for col, val in enumerate(values, 1):
                _data_style(ws, row, col, val, score=(col == 10))

            ws.row_dimensions[row].height = 20
            row += 1

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:M{row - 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_completion_excel(cycle_id: int, db: Session) -> bytes:
    """Which employees/managers have completed check-ins."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Completion Dashboard"

    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = f"AtomQuest — Completion Dashboard | Cycle {cycle_id}"
    title.font = Font(bold=True, size=14, color="1F3864")
    title.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 35

    headers = [
        "Emp ID", "Employee Name", "Department", "Manager",
        "Sheet Status", "Goals Count", "Has Check-in Comment", "Submitted At",
    ]
    _header_style(ws, headers, row=2)

    employees = db.query(User).filter(User.role == "employee", User.is_active == True).all()
    row = 3
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    thin = Side(style="thin", color="EEEEEE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for emp in employees:
        sheet = db.query(GoalSheet).filter(
            GoalSheet.employee_id == emp.id,
            GoalSheet.cycle_id == cycle_id,
        ).first()

        has_checkin = False
        goals_count = 0
        sheet_status = "not_started"
        submitted_at = "—"

        if sheet:
            goals_count = len(sheet.goals)
            sheet_status = sheet.status.value
            has_checkin = len(sheet.checkin_comments) > 0
            submitted_at = sheet.submitted_at.strftime("%Y-%m-%d") if sheet.submitted_at else "—"

        manager_name = emp.manager.name if emp.manager else "—"

        row_data = [
            emp.id, emp.name, emp.department or "—", manager_name,
            sheet_status, goals_count,
            "Yes" if has_checkin else "No",
            submitted_at,
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            # Color-code sheet status
            if col == 5:
                if val == "approved":
                    cell.fill = green
                    cell.font = Font(color="276221")
                elif val in ("submitted", "returned"):
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")
                    cell.font = Font(color="9C5700")
                else:
                    cell.fill = red
                    cell.font = Font(color="9C0006")
            # Color-code check-in
            if col == 7:
                cell.fill = green if val == "Yes" else red

        ws.row_dimensions[row].height = 20
        row += 1

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:H{row - 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
